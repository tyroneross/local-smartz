"""Document processing tools for LangChain agents.

Provides tools for reading PDFs, spreadsheets, and text files.
"""

from langchain_core.tools import tool


@tool
def parse_pdf(file_path: str, pages: str | None = None) -> str:
    """Parse a PDF file and extract its text content. Use pages='1-5' for specific pages."""
    try:
        from pathlib import Path

        path = Path(file_path).expanduser().resolve()
        if not path.exists():
            return f"Error: File not found: {file_path}"

        # Parse page range if specified
        page_indices = None
        if pages:
            try:
                if "-" in pages:
                    start, end = pages.split("-")
                    # Convert 1-indexed to 0-indexed
                    page_indices = list(range(int(start) - 1, int(end)))
                else:
                    page_indices = [int(pages) - 1]
            except (ValueError, IndexError):
                return f"Error: Invalid page range format: {pages}. Use '1-5' or '3'"

        # Try pypdf first for text extraction
        try:
            import pypdf

            with open(path, "rb") as f:
                reader = pypdf.PdfReader(f)
                total_pages = len(reader.pages)

                # Determine which pages to process
                if page_indices:
                    pages_to_read = [i for i in page_indices if 0 <= i < total_pages]
                    if not pages_to_read:
                        return f"Error: Page range {pages} exceeds document length ({total_pages} pages)"
                else:
                    pages_to_read = list(range(total_pages))

                # Extract text
                content_parts = []
                total_chars = 0

                for i in pages_to_read:
                    page = reader.pages[i]
                    text = page.extract_text()
                    total_chars += len(text.strip())
                    content_parts.append(f"## Page {i + 1}\n\n{text.strip()}")

                # Check if text extraction was successful
                avg_chars_per_page = total_chars / len(pages_to_read) if pages_to_read else 0

                if avg_chars_per_page < 100:
                    # Text is too sparse, try pdfplumber for table-heavy PDFs
                    raise ValueError("Sparse text, falling back to pdfplumber")

                result = "\n\n".join(content_parts)
                return f"# PDF: {path.name}\n\nPages: {len(pages_to_read)} of {total_pages}\n\n{result}"

        except (ImportError, ValueError):
            # Fall back to pdfplumber for table-heavy PDFs
            try:
                import pdfplumber

                with pdfplumber.open(path) as pdf:
                    total_pages = len(pdf.pages)

                    # Determine which pages to process
                    if page_indices:
                        pages_to_read = [i for i in page_indices if 0 <= i < total_pages]
                        if not pages_to_read:
                            return f"Error: Page range {pages} exceeds document length ({total_pages} pages)"
                    else:
                        pages_to_read = list(range(total_pages))

                    # Extract text
                    content_parts = []

                    for i in pages_to_read:
                        page = pdf.pages[i]
                        text = page.extract_text() or ""
                        content_parts.append(f"## Page {i + 1}\n\n{text.strip()}")

                    result = "\n\n".join(content_parts)
                    return f"# PDF: {path.name}\n\nPages: {len(pages_to_read)} of {total_pages}\n\n{result}"

            except ImportError:
                return "Error: Neither pypdf nor pdfplumber is installed. Install with: pip install pypdf pdfplumber"

    except Exception as e:
        return f"Error parsing PDF: {type(e).__name__}: {e}"


@tool
def read_spreadsheet(file_path: str, sheet_name: str | None = None, max_rows: int = 100) -> str:
    """Read an Excel spreadsheet and return contents as markdown tables."""
    try:
        from pathlib import Path

        path = Path(file_path).expanduser().resolve()
        if not path.exists():
            return f"Error: File not found: {file_path}"

        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl is not installed. Install with: pip install openpyxl"

        # Load workbook
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            return f"Error loading workbook: {type(e).__name__}: {e}"

        # Get sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                available = ", ".join(wb.sheetnames)
                return f"Error: Sheet '{sheet_name}' not found. Available sheets: {available}"
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        # Read data
        rows_list = list(ws.iter_rows(values_only=True))
        total_rows = len(rows_list)

        if total_rows == 0:
            return f"# Spreadsheet: {path.name}\n\nSheet: {sheet_name}\n\nNo data found."

        # Cap at max_rows
        rows_to_read = min(max_rows, total_rows)
        rows_data = rows_list[:rows_to_read]

        # Format as markdown table
        if not rows_data:
            return f"# Spreadsheet: {path.name}\n\nSheet: {sheet_name}\n\nNo data found."

        # Build table
        lines = []

        # Header row
        header = rows_data[0]
        header_str = "| " + " | ".join(str(cell) if cell is not None else "" for cell in header) + " |"
        lines.append(header_str)

        # Separator
        separator = "| " + " | ".join("---" for _ in header) + " |"
        lines.append(separator)

        # Data rows
        for row in rows_data[1:]:
            row_str = "| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |"
            lines.append(row_str)

        table = "\n".join(lines)

        # Build result
        result_parts = [
            f"# Spreadsheet: {path.name}",
            f"\nSheet: {sheet_name}",
            f"\nAvailable sheets: {', '.join(wb.sheetnames)}",
            f"\nRows read: {rows_to_read} of {total_rows}",
            f"\n\n{table}"
        ]

        if rows_to_read < total_rows:
            result_parts.append(f"\n\n[Showing first {max_rows} rows. Total rows: {total_rows}]")

        return "".join(result_parts)

    except Exception as e:
        return f"Error reading spreadsheet: {type(e).__name__}: {e}"


@tool
def read_text_file(file_path: str, max_lines: int = 500) -> str:
    """Read a text file and return its contents. Caps at max_lines."""
    try:
        from pathlib import Path

        path = Path(file_path).expanduser().resolve()
        if not path.exists():
            return f"Error: File not found: {file_path}"

        # Try common encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        content = None

        for encoding in encodings:
            try:
                with open(path, 'r', encoding=encoding) as f:
                    lines = f.readlines()
                    content = lines
                break
            except (UnicodeDecodeError, LookupError):
                continue

        if content is None:
            return f"Error: Unable to decode file with common encodings: {', '.join(encodings)}"

        total_lines = len(content)

        # Cap at max_lines
        if total_lines > max_lines:
            result = "".join(content[:max_lines])
            result += f"\n\n[Truncated at {max_lines} lines. File has {total_lines} lines total.]"
        else:
            result = "".join(content)

        return result

    except Exception as e:
        return f"Error reading text file: {type(e).__name__}: {e}"
